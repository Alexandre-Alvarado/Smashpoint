from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.db import models
import io
try:
    import qrcode
except ImportError:
    qrcode = None
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.contrib.auth.models import Group, User

from .models import Jugador, Torneo, Resultado, Inscripcion, Partido, Ranking, Grupo
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import openpyxl
from .forms import FormJugador, FormTorneo, FormResultado, FormContacto, FormInscripcion, FormPartido, BulkJugadorImportForm

# Create your views here.

def login_view(request):

    if request.user.is_authenticated:
        return redirect('index')

    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        if user:
            login(request, user)
            return redirect('index')
        else:
            messages.error(request, "Usuario o contraseña incorrectos")

    return render(request, "login.html")


def logout_view(request):
    logout(request)
    return redirect('login')


@login_required
def index(request):
    from datetime import date
    torneos_activos = Torneo.objects.filter(estado__in=['ABIERTO','EN_CURSO']).count()
    jugadores_total = Jugador.objects.count()
    partidos_hoy = Partido.objects.filter(torneo__fecha=date.today()).count()
    top_ranking = Ranking.objects.select_related('jugador').order_by('-puntos')[:5]
    # Último campeón: partido FINAL más reciente con ganador
    ultimo_partido_final = Partido.objects.filter(etapa='FINAL', ganador__isnull=False).select_related('torneo','ganador').order_by('-id').first()
    ultimo_campeon = None
    if ultimo_partido_final:
        ultimo_campeon = {
            'nombre': ultimo_partido_final.ganador.nombre,
            'apellido': ultimo_partido_final.ganador.apellido,
            'torneo': ultimo_partido_final.torneo.nombre
        }
    return render(request, "index_mejorado.html", {
        'torneos_activos': torneos_activos,
        'jugadores_total': jugadores_total,
        'partidos_hoy': partidos_hoy,
        'top_ranking': top_ranking,
        'ultimo_campeon': ultimo_campeon
    })


@login_required
def lista_jugadores(request):
    jugadores = Jugador.objects.all()
    es_admin = request.user.has_perm("smashpointApp.add_jugador")
    return render(request, "jugadores/lista.html", {
        "jugadores": jugadores,
        "es_admin": es_admin
    })


@login_required
@permission_required('smashpointApp.add_jugador', raise_exception=True)
def importar_jugadores_excel(request):
    form = BulkJugadorImportForm(request.POST or None, request.FILES or None)
    reporte = []
    if request.method == 'POST' and form.is_valid():
        archivo = form.cleaned_data['archivo']
        from openpyxl import load_workbook
        try:
            wb = load_workbook(filename=archivo, data_only=True)
        except Exception as e:
            messages.error(request, f"Error leyendo Excel: {e}")
            return render(request, 'jugadores/importar.html', {'form': form})
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        # Normalizar headers
        mapa = {h.lower(): idx for idx, h in enumerate(headers) if h}
        requeridos = ['nombre','apellido','categoria','rut']
        faltan = [r for r in requeridos if r not in mapa]
        if faltan:
            messages.error(request, f"Faltan columnas requeridas: {', '.join(faltan)}")
            return render(request, 'jugadores/importar.html', {'form': form})
        creados = 0
        for row in ws.iter_rows(min_row=2):
            valores = [c.value for c in row]
            nombre = (valores[mapa['nombre']] or '').strip()
            apellido = (valores[mapa['apellido']] or '').strip()
            categoria = (valores[mapa['categoria']] or 'AMATEUR').strip().upper()
            rut = (valores[mapa['rut']] or '').strip()
            if not nombre or not apellido or not rut:
                reporte.append(f"Fila {row[0].row}: datos incompletos, omitida.")
                continue
            # Validar categoría
            if categoria not in ['AMATEUR','FEDERADO']:
                categoria = 'AMATEUR'
            # Verificar duplicado rut
            if Jugador.objects.filter(rut__iexact=rut).exists():
                reporte.append(f"Fila {row[0].row}: RUT duplicado {rut}, omitido.")
                continue
            # Crear jugador (validador rut en modelo puede lanzar excepción)
            try:
                jugador = Jugador(nombre=nombre, apellido=apellido, categoria=categoria, rut=rut)
                jugador.full_clean()
                jugador.save()
                creados += 1
            except Exception as e:
                reporte.append(f"Fila {row[0].row}: error RUT {rut} -> {e}")
        messages.success(request, f"Importación finalizada. Jugadores creados: {creados}")
        if not reporte:
            reporte.append("Sin incidencias.")
    return render(request, 'jugadores/importar.html', {'form': form, 'reporte': reporte})



@login_required
@permission_required('smashpointApp.add_jugador', raise_exception=True)
def agregar_jugador(request):
    form = FormJugador(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect("lista_jugadores")
    return render(request, "jugadores/agregar.html", {"form": form})


@login_required
@permission_required('smashpointApp.change_jugador', raise_exception=True)
def editar_jugador(request, id):
    jugador = get_object_or_404(Jugador, id=id)
    form = FormJugador(request.POST or None, instance=jugador)
    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect("lista_jugadores")
    return render(request, "jugadores/editar.html", {"form": form})


@login_required
@permission_required('smashpointApp.delete_jugador', raise_exception=True)
def eliminar_jugador(request, id):
    jugador = get_object_or_404(Jugador, id=id)
    jugador.delete()
    return redirect("lista_jugadores")


@login_required
def lista_torneos(request):
    torneos = Torneo.objects.all()
    es_admin = request.user.has_perm("smashpointApp.add_torneo")
    return render(request, "torneos/lista.html", {
        "torneos": torneos,
        "es_admin": es_admin
    })


@login_required
@permission_required('smashpointApp.add_partido', raise_exception=True)
def generar_grupos(request, torneo_id):
    torneo = get_object_or_404(Torneo, id=torneo_id)
    if torneo.numero_grupos <= 0:
        messages.error(request, 'El torneo no tiene número de grupos configurado.')
        return redirect('lista_torneos')
    # No regenerar si ya existen grupos
    if Grupo.objects.filter(torneo=torneo).exists():
        return redirect('lista_grupos', torneo_id=torneo.id)
    inscritos_ids = list(Inscripcion.objects.filter(torneo=torneo, estado='INSCRITO').values_list('jugador_id', flat=True))
    jugadores = list(Jugador.objects.filter(id__in=inscritos_ids))
    if len(jugadores) == 0:
        messages.error(request, 'No hay jugadores inscritos.')
        return redirect('lista_torneos')
    
    # Calcular número óptimo de grupos para que el bracket sea potencia de 2
    # Tomamos los 2 mejores de cada grupo, entonces jugadores_bracket = num_grupos * 2
    # Necesitamos que sea potencia de 2 (2, 4, 8, 16, 32, 64, etc.)
    def potencia_2_cercana(n):
        """Encuentra la potencia de 2 más cercana menor o igual a n"""
        import math
        if n <= 2:
            return 2
        return 2 ** math.floor(math.log2(n))
    
    # Calcular cuántos jugadores pueden avanzar (potencia de 2)
    max_bracket_size = potencia_2_cercana(len(jugadores))
    
    # Calcular número de grupos (cada grupo manda 2 jugadores)
    num_grupos = max_bracket_size // 2
    
    # No truncamos jugadores: todos deben jugar grupos; clasificarán los mejores 2 de cada grupo
    jugadores = sorted(jugadores, key=lambda x: x.id)
    tamaño_grupo = (len(jugadores) + num_grupos - 1) // num_grupos  # Tamaño promedio
    
    # Distribuir jugadores en grupos
    grupos_creados = []
    for g_idx in range(num_grupos):
        g = Grupo.objects.create(torneo=torneo, nombre=chr(65 + g_idx))
        grupos_creados.append(g)
    
    # Asignar jugadores de forma equitativa
    for idx, j in enumerate(jugadores):
        g = grupos_creados[idx % num_grupos]
        g.jugadores.add(j)
    
    # Generar partidos: todos contra todos dentro de cada grupo
    for g in grupos_creados:
        grp_players = list(g.jugadores.all())
        for i in range(len(grp_players)):
            for k in range(i+1, len(grp_players)):
                Partido.objects.create(
                    torneo=torneo,
                    ronda=1,
                    etapa='GRUPOS',
                    grupo=g.nombre,
                    jugador_a=grp_players[i],
                    jugador_b=grp_players[k],
                    best_of=3
                )
    torneo.estado = 'EN_CURSO'
    torneo.save()
    bracket_size = num_grupos * 2
    messages.success(request, f'Grupos generados: {num_grupos} grupos. Bracket: {bracket_size} mejores jugadores.')
    return redirect('lista_grupos', torneo_id=torneo.id)


@login_required
def lista_grupos(request, torneo_id):
    torneo = get_object_or_404(Torneo, id=torneo_id)
    grupos = Grupo.objects.filter(torneo=torneo).order_by('nombre')
    data = []
    for g in grupos:
        partidos = Partido.objects.filter(grupo=g.nombre, torneo=torneo, etapa='GRUPOS').order_by('ronda', 'id')
        data.append({
            'grupo': g,
            'tabla': g.estadisticas(),
            'partidos': partidos
        })
    return render(request, 'grupos/lista.html', {'torneo': torneo, 'grupos_data': data})


@login_required
@permission_required('smashpointApp.add_partido', raise_exception=True)
def generar_bracket(request, torneo_id):
    torneo = get_object_or_404(Torneo, id=torneo_id)
    grupos = Grupo.objects.filter(torneo=torneo)
    if not grupos.exists():
        messages.error(request, 'Primero genere los grupos.')
        return redirect('lista_torneos')
    # Verificar todos partidos de grupos con ganador
    if Partido.objects.filter(torneo=torneo, etapa='GRUPOS', ganador__isnull=True).exists():
        messages.error(request, 'Aún hay partidos de grupos sin resultado.')
        return redirect('lista_grupos', torneo_id=torneo.id)
    # Reunir clasificados (top 2 por grupo)
    clasificados = []
    for g in grupos:
        stats = g.estadisticas()
        clasificados.extend([s['jugador'] for s in stats[:2]])
    # Evitar regenerar si ya hay eliminación
    if Partido.objects.filter(torneo=torneo, etapa__in=['ELIMINACION','FINAL']).exists():
        return redirect('lista_partidos', torneo_id=torneo.id)
    # Crear primera ronda eliminación
    ronda = 1
    for i in range(0, len(clasificados), 2):
        if i+1 < len(clasificados):
            Partido.objects.create(
                torneo=torneo,
                ronda=ronda,
                etapa='ELIMINACION',
                jugador_a=clasificados[i],
                jugador_b=clasificados[i+1],
                best_of=3
            )
    messages.success(request, 'Bracket de eliminación generado.')
    return redirect('lista_partidos', torneo_id=torneo.id)


@login_required
@permission_required('smashpointApp.add_partido', raise_exception=True)
def generar_eliminacion_siguiente(request, torneo_id):
    torneo = get_object_or_404(Torneo, id=torneo_id)
    actuales = Partido.objects.filter(torneo=torneo, etapa__in=['ELIMINACION','FINAL']).order_by('ronda')
    if not actuales.exists():
        messages.error(request, 'No hay partidos de eliminación para avanzar.')
        return redirect('lista_partidos', torneo_id=torneo.id)
    max_ronda = actuales.aggregate(m=models.Max('ronda'))['m']
    ronda_partidos = Partido.objects.filter(torneo=torneo, ronda=max_ronda, etapa__in=['ELIMINACION','FINAL'])
    if ronda_partidos.filter(ganador__isnull=True).exists():
        messages.error(request, 'Aún hay partidos sin ganador en la ronda actual.')
        return redirect('lista_partidos', torneo_id=torneo.id)
    ganadores = [p.ganador for p in ronda_partidos]
    if len(ganadores) == 1:
        # Ya hay campeón
        torneo.estado = 'FINALIZADO'
        torneo.save()
        messages.success(request, f'Torneo finalizado. Campeón: {ganadores[0]}')
        return redirect('lista_partidos', torneo_id=torneo.id)
    nueva_ronda = max_ronda + 1
    etapa = 'FINAL' if len(ganadores) == 2 else 'ELIMINACION'
    bo = 5 if etapa == 'FINAL' else 3
    for i in range(0, len(ganadores), 2):
        if i+1 < len(ganadores):
            Partido.objects.create(
                torneo=torneo,
                ronda=nueva_ronda,
                etapa=etapa,
                jugador_a=ganadores[i],
                jugador_b=ganadores[i+1],
                best_of=bo
            )
    messages.success(request, f'Ronda {nueva_ronda} de eliminación generada.')
    return redirect('lista_partidos', torneo_id=torneo.id)



@login_required
@permission_required('smashpointApp.add_torneo', raise_exception=True)
def agregar_torneo(request):
    form = FormTorneo(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect("lista_torneos")
    return render(request, "torneos/agregar.html", {"form": form})


@login_required
@permission_required('smashpointApp.change_torneo', raise_exception=True)
def editar_torneo(request, id):
    torneo = get_object_or_404(Torneo, id=id)
    form = FormTorneo(request.POST or None, instance=torneo)
    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect("lista_torneos")
    return render(request, "torneos/editar.html", {"form": form})


@login_required
@permission_required('smashpointApp.delete_torneo', raise_exception=True)
def eliminar_torneo(request, id):
    torneo = get_object_or_404(Torneo, id=id)
    torneo.delete()
    return redirect("lista_torneos")


# ============================================
# RESULTADOS (DESHABILITADO - NO SE USA)
# ============================================
# Las funciones de resultados han sido deshabilitadas porque
# la gestión de resultados se hace a través de Partidos.
# Se mantiene comentado para referencia.

# @login_required
# def lista_resultados(request):
#     resultados = Resultado.objects.all()
#     es_admin = request.user.has_perm("smashpointApp.add_resultado")
#     return render(request, "resultados/lista.html", {
#         "resultados": resultados,
#         "es_admin": es_admin
#     })

# @login_required
# @permission_required('smashpointApp.add_resultado', raise_exception=True)
# def agregar_resultado(request):
#     form = FormResultado(request.POST or None)
#     if request.method == "POST" and form.is_valid():
#         form.save()
#         return redirect("lista_resultados")
#     return render(request, "resultados/agregar.html", {"form": form})

# @login_required
# @permission_required('smashpointApp.change_resultado', raise_exception=True)
# def editar_resultado(request, id):
#     resultado = get_object_or_404(Resultado, id=id)
#     form = FormResultado(request.POST or None, instance=resultado)
#     if request.method == "POST" and form.is_valid():
#         form.save()
#         return redirect("lista_resultados")
#     return render(request, "resultados/editar.html", {"form": form})

# @login_required
# @permission_required('smashpointApp.delete_resultado', raise_exception=True)
# def eliminar_resultado(request, id):
#     resultado = get_object_or_404(Resultado, id=id)
#     resultado.delete()
#     return redirect("lista_resultados")


# ------------ CONTACTO ------------
@login_required
def contacto(request):
    form = FormContacto(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        return render(request, "gracias.html")  
    return render(request, "contacto.html", {"form": form})


# ------------ INSCRIPCIONES ------------
@login_required
def inscribir_jugador(request, torneo_id):
    torneo = get_object_or_404(Torneo, id=torneo_id)
    jugadores = Jugador.objects.all().order_by('nombre','apellido')
    inscritos_actuales = Inscripcion.objects.filter(torneo=torneo, estado='INSCRITO').count()
    cupos_disponibles = torneo.cupos_max - inscritos_actuales

    if request.method == 'POST':
        ids = request.POST.getlist('jugadores')
        if not ids:
            messages.error(request, 'Selecciona al menos un jugador.')
        else:
            # Contar cuántos nuevos se intentarán inscribir (no contar los ya inscritos)
            nuevos = 0
            creados = 0
            repetidos = 0
            excedentes = []
            
            for jid in ids:
                try:
                    jugador = Jugador.objects.get(id=jid)
                except Jugador.DoesNotExist:
                    continue
                
                # Verificar si ya está inscrito
                if Inscripcion.objects.filter(torneo=torneo, jugador=jugador).exists():
                    repetidos += 1
                else:
                    nuevos += 1
                    if nuevos > cupos_disponibles:
                        excedentes.append(f"{jugador.nombre} {jugador.apellido}")
            
            # Validar que no se exceda cupo
            if nuevos > cupos_disponibles:
                msg = f'Cupos insuficientes. Disponibles: {cupos_disponibles}, intentados: {nuevos}.'
                if excedentes:
                    msg += f' No se pueden inscribir: {", ".join(excedentes[:3])}'
                    if len(excedentes) > 3:
                        msg += f' y {len(excedentes)-3} más.'
                messages.error(request, msg)
            else:
                # Proceder con inscripciones
                for jid in ids:
                    try:
                        jugador = Jugador.objects.get(id=jid)
                    except Jugador.DoesNotExist:
                        continue
                    insc, created = Inscripcion.objects.get_or_create(torneo=torneo, jugador=jugador)
                    if created:
                        creados += 1
                    else:
                        repetidos += 1
                
                if creados:
                    messages.success(request, f'Inscripciones creadas: {creados}.')
                if repetidos:
                    messages.info(request, f'Ya estaban inscritos: {repetidos}.')
                return redirect('lista_inscripciones', torneo_id=torneo.id)

    return render(request, 'inscripciones/inscribir.html', {
        'torneo': torneo, 
        'jugadores': jugadores,
        'cupos_disponibles': cupos_disponibles,
        'inscritos_actuales': inscritos_actuales
    })

@login_required
def lista_inscripciones(request, torneo_id):
    torneo = get_object_or_404(Torneo, id=torneo_id)
    inscripciones = Inscripcion.objects.filter(torneo=torneo)
    return render(request, 'inscripciones/lista.html', {
        'torneo': torneo,
        'inscripciones': inscripciones
    })


# ------------ FIXTURES / PARTIDOS ------------
@login_required
@permission_required('smashpointApp.add_partido', raise_exception=True)
def generar_fixture(request, torneo_id):
    torneo = get_object_or_404(Torneo, id=torneo_id)
    inscritos = list(Inscripcion.objects.filter(torneo=torneo, estado='INSCRITO').values_list('jugador_id', flat=True))
    jugadores = list(Jugador.objects.filter(id__in=inscritos))
    # Si ya existen partidos no regenerar
    if Partido.objects.filter(torneo=torneo).exists():
        return redirect('lista_partidos', torneo_id=torneo.id)
    # Generar emparejamientos simples secuenciales
    for i in range(0, len(jugadores), 2):
        if i+1 < len(jugadores):
            Partido.objects.create(
                torneo=torneo,
                ronda=1,
                jugador_a=jugadores[i],
                jugador_b=jugadores[i+1]
            )
    torneo.estado = 'EN_CURSO'
    torneo.calcular_total_rondas()
    torneo.save()
    return redirect('lista_partidos', torneo_id=torneo.id)

@login_required
def lista_partidos(request, torneo_id):
    torneo = get_object_or_404(Torneo, id=torneo_id)
    
    # Si hay partidos de eliminación o final, mostrar solo esos (descartando grupos)
    if Partido.objects.filter(torneo=torneo, etapa__in=['ELIMINACION', 'FINAL']).exists():
        # Obtener la ronda máxima de eliminación/final
        max_ronda = Partido.objects.filter(torneo=torneo, etapa__in=['ELIMINACION', 'FINAL']).aggregate(
            m=models.Max('ronda')
        )['m']
        # Mostrar solo partidos de la ronda máxima (actual)
        partidos = Partido.objects.filter(torneo=torneo, etapa__in=['ELIMINACION', 'FINAL'], ronda=max_ronda).order_by('id')
    else:
        # Si no hay eliminación, mostrar todos (grupos)
        partidos = Partido.objects.filter(torneo=torneo).order_by('ronda', 'id')
    
    return render(request, 'partidos/lista.html', {
        'torneo': torneo,
        'partidos': partidos
    })

@login_required
def lista_bracket(request, torneo_id):
    torneo = get_object_or_404(Torneo, id=torneo_id)
    partidos = Partido.objects.filter(torneo=torneo, etapa__in=['ELIMINACION','FINAL']).order_by('ronda')
    return render(request, 'bracket/lista.html', {'torneo': torneo, 'partidos': partidos})

@login_required
def bracket_visual(request, torneo_id):
    torneo = get_object_or_404(Torneo, id=torneo_id)
    partidos = Partido.objects.filter(torneo=torneo, etapa__in=['ELIMINACION','FINAL']).order_by('ronda')
    rondas = {}
    for p in partidos:
        rondas.setdefault(p.ronda, []).append(p)
    orden = sorted(rondas.items())
    return render(request, 'bracket/visual.html', {'torneo': torneo, 'rondas': orden})

@login_required
@permission_required('smashpointApp.change_partido', raise_exception=True)
def editar_partido(request, partido_id):
    partido = get_object_or_404(Partido, id=partido_id)
    form = FormPartido(request.POST or None, instance=partido)
    if request.method == 'POST' and form.is_valid():
        form.save()
        # Si hay detalle de sets actualizar sets y ganador
        if partido.detalle_sets:
            partido.parsear_detalle_sets()
        else:
            partido.calcular_ganador()
        return redirect('lista_partidos', torneo_id=partido.torneo.id)
    return render(request, 'partidos/editar.html', {'form': form, 'partido': partido})

@login_required
@permission_required('smashpointApp.add_partido', raise_exception=True)
def generar_ronda_siguiente(request, torneo_id):
    torneo = get_object_or_404(Torneo, id=torneo_id)
    # Obtener la ronda máxima actual
    max_ronda = Partido.objects.filter(torneo=torneo).aggregate(m=models.Max('ronda'))['m'] or 1
    actuales = Partido.objects.filter(torneo=torneo, ronda=max_ronda)
    
    # Verificar que hay partidos en la ronda actual
    if not actuales.exists():
        messages.error(request, 'No hay partidos en la ronda actual para avanzar.')
        return redirect('lista_partidos', torneo_id=torneo.id)
    
    # Verificar que todos tienen ganador
    sin_ganador = actuales.filter(ganador__isnull=True)
    if sin_ganador.exists():
        count = sin_ganador.count()
        messages.error(request, f'Aún hay {count} partido(s) sin ganador. Completa todos antes de avanzar.')
        return redirect('lista_partidos', torneo_id=torneo.id)
    
    ganadores = [p.ganador for p in actuales]
    nueva_ronda = max_ronda + 1
    
    # Determinar la etapa según la cantidad de ganadores
    etapa = 'FINAL' if len(ganadores) == 2 else 'ELIMINACION'
    best_of = 5 if etapa == 'FINAL' else 3
    
    for i in range(0, len(ganadores), 2):
        if i+1 < len(ganadores):
            Partido.objects.create(
                torneo=torneo,
                ronda=nueva_ronda,
                etapa=etapa,
                jugador_a=ganadores[i],
                jugador_b=ganadores[i+1],
                best_of=best_of
            )
    if len(ganadores) == 1:
        torneo.estado = 'FINALIZADO'
        torneo.save()
        messages.success(request, 'Torneo finalizado. Campeón: %s' % ganadores[0])
    else:
        messages.success(request, f'Ronda {nueva_ronda} generada correctamente.')
    return redirect('lista_partidos', torneo_id=torneo.id)


# ------------ RANKING PÚBLICO ------------
# Scoreboard deshabilitado: usaba el modelo Resultado obsoleto.
# Se mantiene solo el ranking actualizado por puntos.

def ranking_public(request):
    ranking = Ranking.objects.select_related('jugador').order_by('-puntos')
    return render(request, 'public/ranking.html', {'ranking': ranking})

def registro_jugador(request):
    if request.method == 'POST':
        username = request.POST.get('username','').strip()
        password = request.POST.get('password','').strip()
        first_name = request.POST.get('first_name','').strip()
        last_name = request.POST.get('last_name','').strip()
        categoria = request.POST.get('categoria','AMATEUR').strip().upper()
        rut = request.POST.get('rut','').strip()
        origen = request.POST.get('origen','').strip()
        
        # Validaciones
        if not username or not password:
            messages.error(request, 'Usuario y contraseña son obligatorios.')
            return render(request, 'public/registro.html')
        
        if User.objects.filter(username__iexact=username).exists():
            messages.error(request, f'El usuario "{username}" ya está registrado. Por favor elige otro nombre de usuario.')
            return render(request, 'public/registro.html')
        
        if rut and Jugador.objects.filter(rut__iexact=rut).exists():
            messages.error(request, 'El RUT ingresado ya está registrado.')
            return render(request, 'public/registro.html')
        
        # Crear usuario
        try:
            user = User.objects.create_user(username=username, password=password, first_name=first_name, last_name=last_name)
            user.is_staff = False
            user.save()
            grupo, _ = Group.objects.get_or_create(name='Jugadores')
            user.groups.add(grupo)
            
            # Crear registro en Jugador
            if categoria not in ['AMATEUR','FEDERADO']:
                categoria = 'AMATEUR'
            
            jugador = Jugador(
                nombre=first_name or username, 
                apellido=last_name or '', 
                categoria=categoria, 
                rut=rut or None, 
                origen=origen or None
            )
            jugador.full_clean()
            jugador.save()
            Ranking.objects.get_or_create(jugador=jugador)
            
            # Login automático
            login(request, user)
            messages.success(request, f'¡Bienvenido {username}! Tu cuenta ha sido creada exitosamente.')
            return redirect('ranking_public')
            
        except Exception as e:
            # Si falla algo, eliminar el usuario creado
            if 'user' in locals():
                user.delete()
            messages.error(request, f'Error al crear la cuenta: {str(e)}')
            return render(request, 'public/registro.html')
    
    return render(request, 'public/registro.html')

def jugador_public(request, jugador_id):
    jugador = get_object_or_404(Jugador, id=jugador_id)
    partidos = Resultado.objects.filter(jugador1=jugador) | Resultado.objects.filter(jugador2=jugador)
    try:
        rk = Ranking.objects.get(jugador=jugador)
    except Ranking.DoesNotExist:
        rk = None
    return render(request, 'public/jugador.html', {
        'jugador': jugador,
        'partidos': partidos,
        'ranking': rk
    })

def jugador_qr(request, jugador_id):
    jugador = get_object_or_404(Jugador, id=jugador_id)
    if not qrcode:
        return HttpResponse('QR library not installed', status=500)
    url = request.build_absolute_uri(f'/public/jugador/{jugador.id}/')
    img = qrcode.make(url)
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    return HttpResponse(buf.getvalue(), content_type='image/png')


# ------------ EXPORTACIÓN CSV ------------
import csv

@login_required
def export_jugadores_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="jugadores.csv"'
    writer = csv.writer(response)
    writer.writerow(['id','nombre','apellido','categoria','rut','origen'])
    for j in Jugador.objects.all():
        writer.writerow([j.id,j.nombre,j.apellido,j.categoria,j.rut,j.origen])
    return response

@login_required
def export_torneos_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="torneos.csv"'
    writer = csv.writer(response)
    writer.writerow(['id','nombre','direccion','fecha','categoria','cupos_max','estado'])
    for t in Torneo.objects.all():
        writer.writerow([t.id,t.nombre,t.direccion,t.fecha,t.categoria,t.cupos_max,t.estado])
    return response

@login_required
def export_resultados_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="resultados.csv"'
    writer = csv.writer(response)
    writer.writerow(['id','torneo','j1','j2','m_j1','m_j2'])
    for r in Resultado.objects.all():
        writer.writerow([r.id,r.torneo_id,r.jugador1_id,r.jugador2_id,r.marcador_j1,r.marcador_j2])
    return response

@login_required
def import_jugadores_csv(request):
    if request.method == 'POST' and request.FILES.get('file'):
        f = request.FILES['file']
        decoded = f.read().decode('utf-8').splitlines()
        reader = csv.DictReader(decoded)
        creados = 0
        for row in reader:
            lic = row.get('licencia') or row.get('license')
            if lic and not Jugador.objects.filter(licencia=lic).exists():
                Jugador.objects.create(
                    nombre=row.get('nombre',''),
                    apellido=row.get('apellido',''),
                    categoria=row.get('categoria','AMATEUR'),
                    licencia=lic
                )
                creados += 1
        messages.success(request, f"Importación completa. Jugadores creados: {creados}")
        return redirect('lista_jugadores')
    return render(request, 'import/jugadores.html')


# ------------ API JSON SIMPLE ------------
def api_jugadores(request):
    data = list(Jugador.objects.values('id','nombre','apellido','categoria','rut','origen'))
    return JsonResponse({'jugadores': data})

def api_torneos(request):
    data = list(Torneo.objects.values('id','nombre','fecha','categoria','estado','cupos_max'))
    return JsonResponse({'torneos': data})

def api_ranking(request):
    data = list(Ranking.objects.select_related('jugador').order_by('-puntos').values('jugador__nombre','jugador__apellido','puntos'))
    return JsonResponse({'ranking': data})


# ------------ OFFLINE / MANIFEST ------------
def offline(request):
    return render(request, 'offline.html')


# ------------ EXPORT RANKING PDF / EXCEL ------------
@login_required
def export_ranking_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="ranking.pdf"'
    c = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    c.setFont("Helvetica-Bold", 16)
    c.drawString(50, height - 50, "Ranking SmashPoint")
    c.setFont("Helvetica", 12)
    y = height - 90
    for rk in Ranking.objects.select_related('jugador').order_by('-puntos'):
        c.drawString(50, y, f"{rk.jugador.nombre} {rk.jugador.apellido} - {rk.puntos} pts")
        y -= 18
        if y < 60:
            c.showPage()
            c.setFont("Helvetica", 12)
            y = height - 50
    c.save()
    return response

@login_required
def export_ranking_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ranking'
    ws.append(['Nombre','Apellido','Puntos'])
    for rk in Ranking.objects.select_related('jugador').order_by('-puntos'):
        ws.append([rk.jugador.nombre, rk.jugador.apellido, rk.puntos])
    from openpyxl.utils import get_column_letter
    for col in range(1,4):
        ws.column_dimensions[get_column_letter(col)].width = 20
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="ranking.xlsx"'
    wb.save(response)
    return response
